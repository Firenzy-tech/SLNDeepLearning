import pandas as pd
import streamlit as st
from typing import Tuple, Dict, Any, List, Generator
import io
import numpy as np
import os
import csv
import json
import ijson
import duckdb
from openpyxl import load_workbook
try:
    # Streamlit's UploadedFile type may be in different locations across versions
    from streamlit.runtime.uploaded_file_manager import UploadedFile
except Exception:
    UploadedFile = object
import tempfile

class UniversalFileReader:
    """
    Lector universal para archivos grandes utilizando streaming y procesamiento por lotes.
    Ideal para archivos que superan la memoria RAM.
    """
    def __init__(self, chunk_size=10000):
        self.chunk_size = chunk_size

    def read(self, file_path) -> Generator[List[Dict], None, None]:
        """
        Detecta automáticamente el tipo de archivo y retorna un generador iterable.
        """
        extension = os.path.splitext(file_path)[1].lower()
        handlers = {
            ".csv": self._read_csv,
            ".json": self._read_json,
            ".jsonl": self._read_jsonl,
            ".xlsx": self._read_excel
        }
        if extension not in handlers:
            raise ValueError(f"Formato no soportado: {extension}")
        return handlers[extension](file_path)

    def _read_csv(self, file_path):
        with open(file_path, newline='', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            batch = []
            for row in reader:
                batch.append(row)
                if len(batch) >= self.chunk_size:
                    yield batch
                    batch = []
            if batch:
                yield batch

    def _read_json(self, file_path):
        with open(file_path, "rb") as file:
            parser = ijson.items(file, "item")
            batch = []
            for obj in parser:
                batch.append(obj)
                if len(batch) >= self.chunk_size:
                    yield batch
                    batch = []
            if batch:
                yield batch

    def _read_jsonl(self, file_path):
        with open(file_path, "r", encoding="utf-8") as file:
            batch = []
            for line in file:
                obj = json.loads(line)
                batch.append(obj)
                if len(batch) >= self.chunk_size:
                    yield batch
                    batch = []
            if batch:
                yield batch

    def _read_excel(self, file_path):
        workbook = load_workbook(
            filename=file_path,
            read_only=True,
            data_only=True
        )
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows)
        batch = []
        for row in rows:
            obj = dict(zip(headers, row))
            batch.append(obj)
            if len(batch) >= self.chunk_size:
                yield batch
                batch = []
        if batch:
            yield batch

    def query_csv(self, file_path, query):
        """
        Ejecuta una consulta SQL sobre un archivo CSV usando DuckDB (extremadamente rápido).
        """
        sql = f"""
        SELECT *
        FROM read_csv_auto('{file_path}')
        WHERE {query}
        """
        return duckdb.sql(sql).fetchall()

@st.cache_data(show_spinner=False, hash_funcs={io.BytesIO: lambda _: None, UploadedFile: lambda _: None})
def load_data(uploaded_file, max_rows: int = None, file_id: str = "") -> pd.DataFrame:
    """
    Carga el dataset dependiendo de su extensión.
    Optimizamos el hashing para evitar que Streamlit lea el archivo gigante dos veces.
    """
    try:
        # Reposicionar el puntero del archivo por si Streamlit lo leyó antes
        uploaded_file.seek(0)
        
        suffix = os.path.splitext(uploaded_file.name)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            # Escribir en trozos para no duplicar la memoria RAM con .getbuffer()
            chunk_size = 1024 * 1024 # 1MB chunks
            while True:
                chunk = uploaded_file.read(chunk_size)
                if not chunk: break
                tmp_file.write(chunk)
            tmp_path = tmp_file.name

        try:
            # Para CSV, usamos DuckDB que es mucho más eficiente en memoria
            if suffix == '.csv':
                if max_rows:
                    # Muestreo aleatorio si el archivo es gigante, o primeras N filas
                    query = f"SELECT * FROM read_csv_auto('{tmp_path}') LIMIT {max_rows}"
                    df = duckdb.query(query).to_df()
                else:
                    df = duckdb.query(f"SELECT * FROM read_csv_auto('{tmp_path}')").to_df()
            else:
                # Para otros formatos, usamos el streaming batch a batch para no saturar
                reader = UniversalFileReader(chunk_size=5000)
                all_batches = []
                rows_accumulated = 0
                for batch in reader.read(tmp_path):
                    batch_df = pd.DataFrame(batch)
                    if max_rows:
                        remaining = max_rows - rows_accumulated
                        if remaining <= 0: break
                        if len(batch_df) > remaining: batch_df = batch_df.iloc[:remaining]
                    all_batches.append(batch_df)
                    rows_accumulated += len(batch_df)
                    if max_rows and rows_accumulated >= max_rows: break
                df = pd.concat(all_batches, ignore_index=True)

        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

        return optimize_memory(df)
    except Exception as e:
        raise RuntimeError(f"Error al cargar el archivo: {str(e)}")

def optimize_memory(df: pd.DataFrame) -> pd.DataFrame:
    """
    Optimiza el uso de memoria RAM reduciendo la precisión de los tipos de datos.
    """
    # Use pandas type checks to avoid invalid comparisons (e.g., datetime vs float)
    from pandas.api import types as ptypes

    for col in df.columns:
        try:
            col_series = df[col]
            if ptypes.is_integer_dtype(col_series):
                c_min = col_series.min()
                c_max = col_series.max()
                if c_min >= np.iinfo(np.int8).min and c_max <= np.iinfo(np.int8).max:
                    df[col] = col_series.astype(np.int8)
                elif c_min >= np.iinfo(np.int16).min and c_max <= np.iinfo(np.int16).max:
                    df[col] = col_series.astype(np.int16)
                elif c_min >= np.iinfo(np.int32).min and c_max <= np.iinfo(np.int32).max:
                    df[col] = col_series.astype(np.int32)

            elif ptypes.is_float_dtype(col_series):
                c_min = col_series.min()
                c_max = col_series.max()
                if c_min >= np.finfo(np.float32).min and c_max <= np.finfo(np.float32).max:
                    df[col] = col_series.astype(np.float32)

            elif ptypes.is_bool_dtype(col_series):
                df[col] = col_series.astype(np.uint8)

            elif ptypes.is_categorical_dtype(col_series):
                # already categorical
                continue

            elif ptypes.is_datetime64_any_dtype(col_series):
                # Leave datetime types as-is but ensure dtype is datetime64[ns]
                df[col] = pd.to_datetime(col_series, errors='coerce')

            else:
                # object or mixed types: if low cardinality, convert to category
                try:
                    unique_ratio = len(col_series.unique()) / len(col_series) if len(col_series) > 0 else 0
                except Exception:
                    unique_ratio = 1
                if unique_ratio < 0.5:
                    df[col] = col_series.astype('category')
        except Exception:
            # If any column-specific optimization fails, skip it to keep loading robust
            continue
                
    return df

def get_profile_stats(df: pd.DataFrame) -> Dict[str, Any]:
    """Genera estadísticas descriptivas para el perfilado del dataset."""
    # deep=False es CRÍTICO para archivos grandes, de lo contrario se congela
    stats = {
        "rows": df.shape[0],
        "cols": df.shape[1],
        "missing_cells": df.isnull().sum().sum(),
        "duplicates": df.duplicated().sum(),
        "memory_usage": df.memory_usage(deep=False).sum() / 1024**2 # en MB
    }
    return stats


def get_cleaning_recommendation(dtype: str, null_count: int, total_rows: int, col_name: str) -> str:
    """
    Genera recomendaciones de limpieza según el tipo de dato y características.
    Sigue estándares de ciencia de datos y buenas prácticas.
    """
    null_percentage = (null_count / total_rows * 100) if total_rows > 0 else 0
    
    # Recomendaciones por tipo de dato
    if 'int' in dtype or 'float' in dtype:
        if null_count > 0:
            if null_percentage < 5:
                return f"✅ Imputar con media/mediana (nulos: {null_percentage:.1f}%)"
            elif null_percentage < 20:
                return f"⚠️ Imputar o eliminar filas (nulos: {null_percentage:.1f}%)"
            else:
                return f"❌ Considerar eliminar columna (nulos: {null_percentage:.1f}%)"
        return "✅ Usar escaling (Min-Max o Standardización)"
    
    elif 'object' in dtype or 'string' in dtype:
        if null_count > 0:
            if null_percentage < 5:
                return f"✅ Usar moda o crear categoría 'Unknown' (nulos: {null_percentage:.1f}%)"
            elif null_percentage < 20:
                return f"⚠️ Evaluar imputación o eliminar (nulos: {null_percentage:.1f}%)"
            else:
                return f"❌ Considerar eliminar columna (nulos: {null_percentage:.1f}%)"
        return "✅ Codificar (Label Encoding o One-Hot)"
    
    elif 'datetime' in dtype or 'date' in dtype:
        if null_count > 0:
            return f"✅ Eliminar o interpolar temporalmente (nulos: {null_percentage:.1f}%)"
        return "✅ Extraer features (año, mes, día, día semana)"
    
    elif 'bool' in dtype or 'category' in dtype:
        if null_count > 0:
            return f"⚠️ Imputar con moda (nulos: {null_percentage:.1f}%)"
        return "✅ Codificar como 0/1 o usar One-Hot si hay múltiples categorías"
    
    else:
        return "⚠️ Revisar tipo de dato - considerar conversión"


def get_data_types_with_recommendations(df: pd.DataFrame) -> pd.DataFrame:
    """
    Retorna un DataFrame con recomendaciones de limpieza para cada columna.
    Incluye: Campo, Tipo, Valores Nulos, % Nulos, Recomendación
    """
    total_rows = len(df)
    data_info = []
    
    for col in df.columns:
        null_count = df[col].isnull().sum()
        null_percentage = (null_count / total_rows * 100) if total_rows > 0 else 0
        dtype = str(df[col].dtype)
        recommendation = get_cleaning_recommendation(dtype, null_count, total_rows, col)
        
        data_info.append({
            "📋 Campo": col,
            "🔢 Tipo": dtype,
            "❌ Nulos": null_count,
            "% Nulos": f"{null_percentage:.1f}%",
            "💡 Recomendación": recommendation
        })
    
    return pd.DataFrame(data_info)


